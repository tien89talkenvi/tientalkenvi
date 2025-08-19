import streamlit as st  # streamlit=1.47.1
import pandas as pd     # pandas=2.3.1
import os, time, json, random

from selenium import webdriver  # selenium=4.34.2
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

import shutil
from openpyxl import Workbook, load_workbook    # openpyxl=3.1.5
from openpyxl.styles import PatternFill
from io import BytesIO
import xlsxwriter   # xlsxwriter=3.2.5
import tempfile
from datetime import datetime
import matplotlib.pyplot as plt
import numpy as np


import socket

start_time = time.time()  # Bắt đầu tính giờ

#from playwright.sync_api import sync_playwright    #playwright==1.54.0
#--------------------------------------------------------------

# Cac Ham Phu --------------------------------------------------
# Ham thu viec doc file txt/cvs dung encoding nao khong gay loi
def check_read_file_txt(filetxt):
    encodings_to_try = ['utf-8', 'utf-8-sig', 'cp1252', 'cp1258', 'utf-16']

    for enc in encodings_to_try:
        try:
            df = pd.read_csv('file.txt', delimiter='\t', encoding=enc)
            print(f"✅ Thành công với encoding: {enc}")
            break
        except Exception as e:
            print(f"❌ {enc}: {e}")

# Hàm kiểm tra giá trị có phải số hoặc ngày không
def is_number_or_date(val):
    if pd.isna(val):  # NaN thì giữ lại
        return False
    # Trường hợp là số thật
    if isinstance(val, (int, float)):
        return True
    # Nếu là chuỗi
    if isinstance(val, str):
        # Nếu chuỗi toàn số hoặc dạng số thập phân
        if val.strip().replace('.', '', 1).isdigit():
            return True
        # Thử parse sang datetime
        try:
            datetime.strptime(val.strip(), '%Y-%m-%d')
            return True
        except ValueError:
            pass
        try:
            datetime.strptime(val.strip(), '%d/%m/%Y')
            return True
        except ValueError:
            pass
        try:
            datetime.strptime(val.strip(), '%m-%d-%y')
            return True
        except ValueError:
            pass
    return False

# Cac Ham Cho Phan III --------------------------------------------------

def Ht_Data_tquat(outputIo):
    st.write('Ht_Data_tquat')
    wb = load_workbook(outputIo)
    ws = wb.active
    # xu li wb, ws o day
    # roi save ra file excel de xem
    # Ghi vào memory (không ghi ra ổ đĩa)
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    
    st.download_button(
        label="Tải file Excel về xem",
        data=virtual_workbook.getvalue(),
        file_name="Data_0.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return 

def Ht_Data_sxep(outputIo):
    st.write('Ht_Data_sxep')
    wb = load_workbook(outputIo)
    ws = wb.active
    # xu li wb, ws o day
    # roi save ra file excel de xem
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    
    st.download_button(
        label="Tải file Excel về xem",
        data=virtual_workbook.getvalue(),
        file_name="Data_0.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return 

def Ht_Data_new(outputIo):
    st.write('Ht_Data_new')
    wb = load_workbook(outputIo)
    ws = wb.active
    # xu li wb, ws o day
    # roi save ra file excel de xem
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    
    st.download_button(
        label="Tải file Excel về xem",
        data=virtual_workbook.getvalue(),
        file_name="Data_0.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return 

def Ht_Data_old(outputIo):
    st.write('Ht_Data_old')
    wb = load_workbook(outputIo)
    ws = wb.active
    # xu li wb, ws o day
    # roi save ra file excel de xem
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    
    st.download_button(
        label="Tải file Excel về xem",
        data=virtual_workbook.getvalue(),
        file_name="Data_0.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return 

def Ht_Data_max(outputIo):
    st.write('Ht_Data_max')
    wb = load_workbook(outputIo)
    ws = wb.active
    # xu li wb, ws o day
    # roi save ra file excel de xem
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    
    st.download_button(
        label="Tải file Excel về xem",
        data=virtual_workbook.getvalue(),
        file_name="Data_0.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return 




# Ham tai file txt du lieu dang cvs cua cac mien thuoc bang Cali
@st.cache_data
def download_data_smarts(regions):
    #xoa thu muc downloads va tao lai de chi chua 2 file du lieu
    folder_path_cu = 'downloads'
    # Xóa thư mục nếu tồn tại
    if os.path.exists(folder_path_cu):
        shutil.rmtree(folder_path_cu)  # Xóa toàn bộ thư mục và nội dung bên trong

    download_dir = os.path.abspath("downloads")
    os.makedirs(download_dir, exist_ok=True)

    # ✅ CẤU HÌNH CHROME:
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1   # DÒNG QUAN TRỌNG DE TAT THONG BAO
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--headless")  # chạy ẩn trình duyệt

    # ✅ KHỞI TẠO TRÌNH DUYỆT
    driver = webdriver.Chrome(options=options)

    driver.get("https://smarts.waterboards.ca.gov/smarts/SwPublicUserMenu.xhtml")
    print("✅ Vào trang chính")

    WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.LINK_TEXT, "Download NOI Data By Regional Board"))
    ).click()

    WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
    driver.switch_to.window(driver.window_handles[-1])
    print("✅ Đã chuyển sang tab mới")

    links = [
        "Industrial Application Specific Data",
        "Industrial Ad Hoc Reports - Parameter Data",
        "Industrial Annual Reports"
    ]

    def wait_for_download_and_get_new_file(before_files, timeout=40):
        for _ in range(timeout * 2):
            time.sleep(0.5)
            after_files = set(os.listdir(download_dir))
            new_files = after_files - before_files
            txt_files = [f for f in new_files if f.endswith(".txt")]
            if txt_files:
                return txt_files[0]
        return None
    #---------------------
    region = regions
    print(f"\n🔹 Chọn Region: {region}")
    dropdown = WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.NAME, "intDataFileDowloaddataFileForm:intDataDumpSelectOne"))
    )
    Select(dropdown).select_by_visible_text(region)
    time.sleep(3)  # Đợi dropdown load lại
    
    lfile_datai = []

    for j, name in enumerate(links):
        try:
            print(f"📥 Đang click tải: {name}")
            before = set(os.listdir(download_dir))

            link_elem = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.LINK_TEXT, name))
            )
            driver.execute_script("arguments[0].click();", link_elem)

            fname = wait_for_download_and_get_new_file(before)
            if fname:
                # Tạo tên file chuẩn theo Region + tên file
                src = os.path.join(download_dir, fname)
                dst_name = f"{region} - {name}.txt"
                dst_name = dst_name.replace(" ", "_")  # Nếu muốn
                dst = os.path.join(download_dir, dst_name)
                os.rename(src, dst)
                print(f"File đã lưu: {dst}")
                lfile_datai.append(f"{dst}")
            else:
                print("❌ Không tìm thấy file mới sau khi tải")
        except Exception as e:
            print(f"❌ Lỗi khi tải {name} ở Region {region}: {e}")

    driver.quit()
    print("\n🎉 Hoàn tất tải file cho "+region)
    return lfile_datai
    # CHU Y rang neu ten file dat trung voi file da co thi that bai.

# CAC HAM CHINH-----------------------------------------------------
def ThucThiPhan_4():
    return    


@st.cache_data
def Doc_hthi_data(uploaded_file):
    try:
        # Đọc file Excel thành DataFrame
        df = pd.read_excel(uploaded_file, sheet_name='Data')

        # 1. Sắp xếp dữ liệu theo nhiều cấp độ (multi-level sort):
        df_sorted = df.sort_values(
                by=["OLD/NEW", "PARAMETER", "RESULT"],
                ascending=[True, True, False]
        )
        # 2. Lọc dữ liệu có OLD/NEW == 'New':
        df_new = df_sorted[df_sorted["OLD/NEW"] == "New"]
        # 3. Tô màu (highlight) exceedances thì không thể hiển thị trong DataFrame thông thường nhưng có thể dùng:
        # pandas.ExcelWriter + openpyxl để ghi file Excel có màu.
        # Hoặc đơn giản chỉ đánh dấu bằng cột mới "Exceed" = True/False
        # 4. So sánh kết quả với ngưỡng NAL/NEL/TNAL:
        # tao dic chua nguong
        nal_thresholds = {
            "Ammonia": 4.7,
            "Cadmium": 0.0031,
            "Copper": 0.06749,
            # v.v...
        }
        # Rồi kiểm tra:
        def is_exceed(row):
            param = row["PARAMETER"]
            result = row["RESULT"]
            return result > nal_thresholds.get(param, float('inf'))
        
        df_new["EXCEED"] = df_new.apply(is_exceed, axis=1)
        # 5. Ghi chú các facility cần theo dõi → bạn có thể lọc hoặc thêm cột "Flagged" dựa vào danh sách thủ công.


        st.success(f"Đã tải lên: {uploaded_file.name}")
        st.subheader("📄 Dữ liệu từ file:")

        # Bước 3: Hiển thị DataFrame với cuộn dọc (giả lập 3 dòng)
        st.dataframe(df) #, height=120, use_container_width=True)

    except Exception as e:
        st.error(f"Lỗi khi đọc file: {e}")
    
def Ht_CaiMuonXem_0(tepxlsx):
    # 1.Đọc file Excel thành DataFrame
    df = pd.read_excel(tepxlsx, sheet_name='Data')
    # 2. Sắp xếp dữ liệu theo nhiều cấp độ (multi-level sort):
    df_sorted = df.sort_values(
        by=["OLD/NEW", "PARAMETER", "RESULT"],
        ascending=[True, True, False]
    )
    #Dua vao xlsx de xem kq sx
    
    tepxlsxdexem = "Data_Tracker_3-1.xlsx"
    df_sorted.to_excel(tepxlsxdexem, sheet_name='Datanew', index=False)
    os.startfile(tepxlsxdexem)

#---------------------------------
def Ht_CaiMuonXem_1(uploaded_file):
    uploaded_file = st.file_uploader("Tải lên Data_Tracker_New.xlsx", type=["xlsx"])
    if uploaded_file:
        df = pd.read_excel(uploaded_file)
        # Ghi tạm ra file Excel để xử lý openpyxl
        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        # Load và tô màu
        wb = load_workbook(output)
        ws = wb.active

        # Tìm vị trí các cột "OLD/NEW" và "COSO"
        header = [cell.value for cell in ws[1]]

        try:
            old_new_col_idx = header.index("OLD/NEW") + 1
            coso_col_idx = header.index("FACILITY_NAME") + 1
        except ValueError as e:
            raise Exception(f"Không tìm thấy cột: {e}")

        # Tô màu vàng cho dòng 'new' thuộc cơ sở 'CS1'
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Duyệt từng dòng
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            old_new_val = str(row[old_new_col_idx - 1].value).strip().lower() if row[old_new_col_idx - 1].value else ""
            coso_val = str(row[coso_col_idx - 1].value).strip() if row[coso_col_idx - 1].value else ""

            if old_new_val == "old" and coso_val == 'CS1':
                for cell in row:
                    cell.fill = yellow_fill
                # Dòng này được giữ lại
            else:
                # Ẩn dòng không khớp điều kiện
                ws.row_dimensions[row[0].row].hidden = True

        # Lưu file mới
        # Ghi vào memory (không ghi ra ổ đĩa)
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)

        st.download_button(
            label="Tải file Excel",
            data=virtual_workbook.getvalue(),
            file_name="data_tracker_1.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        #wb.save("filtered_CS1_new.xlsx")
        #st.write('filtered_CS1_new.xlsx da co')        
        #os.startfile("filtered_CS1_new.xlsx")

        # Lưu lại và cho phép tải xuống
        #final_output = BytesIO()
        #wb.save(final_output)
        #final_output.seek(0)

        #st.download_button("📥 Tải file đã tô màu", final_output, "T_to_mau.xlsx")


# ---- phuc vu cho ThucThiPhan_2()------------------
@st.cache_data
def Combined_to_data_tracker(uploaded_tracker):
    try:
        # Read input files
        sheet1 = pd.read_excel(uploaded_tracker, sheet_name='Sheet1', header=None, dtype=str)
        sheet2 = pd.read_excel(uploaded_tracker, sheet_name='Sheet2', dtype=str)
        df_tracker = pd.read_excel(uploaded_tracker, sheet_name='Data', dtype=str)

        # Clean Sheet1
        cols_to_delete = list(range(15, 28)) + list(range(31, 38)) + [0, 4, 5, 6, 7, 11]
        # Chỉ lấy các chỉ số cột cần xóa, nhưng phải nhỏ hơn num_cols
        num_cols = sheet1.shape[1]
        cols_to_delete = [col for col in cols_to_delete if col < num_cols]
        sheet1_cleaned = sheet1.drop(sheet1.columns[cols_to_delete], axis=1)

        sheet1_cleaned.columns = sheet1_cleaned.iloc[0]
        sheet1_cleaned = sheet1_cleaned[1:]

        # Reorder columns
        cols = list(sheet1_cleaned.columns)
        if 'WDID' in cols and 'APP_ID' in cols:
            cols.remove('WDID')
            cols.insert(cols.index('APP_ID'), 'WDID')
        if 'FACILITY_NAME' in cols and 'OPERATOR_NAME' in cols:
            cols.remove('FACILITY_NAME')
            cols.insert(cols.index('OPERATOR_NAME'), 'FACILITY_NAME')
        sheet1_cleaned = sheet1_cleaned[cols]


        # Clean Sheet2 and merge SIC data
        sheet2_cleaned = sheet2.drop(sheet2.columns[2:8], axis=1)
        sheet2_cleaned = sheet2_cleaned.iloc[:, :5]  # Chỉ lấy 5 cột đầu tiên
        sheet2_cleaned.columns = ['A', 'APP_ID', 'PRIMARY_SIC', 'SECONDARY_SIC', 'TERTIARY_SIC']
        sheet2_cleaned = sheet2_cleaned[['APP_ID', 'PRIMARY_SIC', 'SECONDARY_SIC', 'TERTIARY_SIC']]
        merged = pd.merge(sheet1_cleaned, sheet2_cleaned, on='APP_ID', how='left')
        merged = merged.rename(columns={'PRIMARY_SIC': '1', 'SECONDARY_SIC': '2', 'TERTIARY_SIC': '3'})
        merged['2'] = merged['2'].replace('0', pd.NA)
        merged['3'] = merged['3'].replace('0', pd.NA)
        merged['OLD/NEW'] = 'new'

        # Add "Old" tag to tracker and combine
        df_tracker['OLD/NEW'] = 'old'
        df_combined = pd.concat([df_tracker, merged], ignore_index=True)

        # Export to Excel in-memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_combined.to_excel(writer, sheet_name='Data', index=False)
        output.seek(0)
        return output 
    except Exception as e:
        st.error(f"⚠️ An error occurred: {e}")
#----------------

@st.cache_data
def Txt_to_data_tracker(df1, df2, df_data):
    # tao file excel ao xlsx_ao_chua_3df chua df1, df2, df_data
    xlsx_ao_chua_3df = BytesIO()
    with pd.ExcelWriter(xlsx_ao_chua_3df, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
        df_data.to_excel(writer, sheet_name="Data", index=False)

    xlsx_ao_chua_3df.seek(0)
    return xlsx_ao_chua_3df
    # tao nut download file tren neu can
    #st.download_button("📥 Download Data_Tracker_include_3df.xlsx",
    #    data=xlsx_ao_chua_3df,
    #    file_name="Data_Tracker_include_3df.xlsx",
    #    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #)

# Cac ham Phan II -----------------------------    
# ham nay de In Sheet2, Tìm các dòng APP_ID trùng, 
# rồi xóa các dòng mà STATUS ≠ Active nhưng giữ lại các dòng có STATUS = Active "
def Xli_P2_1(F_excel_data_ao):
    F_excel_data_ao.seek(0)  # quay lại đầu BytesIO để đọc
    dfSheet2 = pd.read_excel(F_excel_data_ao, sheet_name='Sheet2')
    # Tìm APP_ID bị trùng
    duplicates = dfSheet2[dfSheet2.duplicated(subset='APP_ID', keep=False)]
    # Giữ lại dòng trùng có STATUS khác "Active"
    to_delete = duplicates[duplicates['STATUS'] != 'Active']
    # Xoá các dòng này khỏi dataframe gốc (chu y la file excel van y cu)
    dfSheet2_cleaned = dfSheet2.drop(to_delete.index)
    # ket qua la cac dong trung APP_ID nhung co STATUS la Active duoc giu lai, con
    # cac dong trung APP_ID nhung co STATUS khac Active thi bi xoa
    # cho hien thi df con lai sau khi da xoa cac to_delete
    # phai hieu df_cleaned la df con lai sau khi da lam sach 
    #st.write(df_cleaned)
    #---
    if st.checkbox("View Sheet2_1", key='BCB1'):
        st.write('(rows, cols) = ', len(dfSheet2_cleaned), len(dfSheet2_cleaned.columns))
        st.write(dfSheet2_cleaned)

    # Ghi thêm dfnew vào Sheet2 mà không mất Sheet1
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet2_cleaned.to_excel(writer, sheet_name="Sheet2", index=False)
    # Giờ output có cả Sheet1 và Sheet2
    F_excel_data_ao.seek(0)
    return F_excel_data_ao

# ham nay de Delete, move, re-order columns in Sheet2
def Xli_P2_2(F_excel_data_ao):
    # set dfSheet2 is dataframe of  Sheet2 in uploaded_f3
    dfSheet2 = pd.read_excel(F_excel_data_ao, sheet_name='Sheet2')
    # xoa cac cot AF den AL: 
    # loai bo cac cot trong [] va tra ve kq dfSheet2_cleaned la data con lai
    dfSheet2_cleaned = dfSheet2.drop(['RECEIVING_WATER_NAME','INDIRECTLY',
        'DIRECTLY','CERTIFIER_BY','CERTIFIER_TITLE','CERTIFICATION_DATE',
        'QUESTION_TMDL_ANSWER'],axis=1)
    # tiep tuc loai bo cac cot P den AB:
    dfSheet2_cleaned = dfSheet2_cleaned.drop(['FACILITY_LATITUDE','FACILITY_LONGITUDE',
    	'FACILITY_COUNTY','FACILITY_CONTACT_FIRST_NAME','FACILITY_CONTACT_LAST_NAME',
        'FACILITY_TITLE','FACILITY_PHONE','FACILITY_EMAIL',	'FACILITY_TOTAL_SIZE',
        'FACILITY_TOTAL_SIZE_UNIT',	'FACILITY_AREA_ACTIVITY','FACILITY_AREA_ACTIVITY_UNIT',
        'PERCENT_OF_SITE_IMPERVIOUSNESS'],axis=1)
    # tiep tuc loai bo cac cot A,E-H,L:
    dfSheet2_cleaned = dfSheet2_cleaned.drop(['PERMIT_TYPE','NOI_PROCESSED_DATE',
        'NOT_EFFECTIVE_DATE','REGION_BOARD','COUNTY', 'FACILITY_ADDRESS_2'],axis=1)
    # di chuyen cot WDID sang ben trai cot APP_ID
    cols = list(dfSheet2_cleaned.columns)
    if 'WDID' in cols and 'APP_ID' in cols:
        cols.remove('WDID')
        cols.insert(cols.index('APP_ID'), 'WDID')
    # tiep tuc di chuyen cot FACILITY_NAME sang ben trai cot OPERATOR_NAME
    if 'FACILITY_NAME' in cols and 'OPERATOR_NAME' in cols:
        cols.remove('FACILITY_NAME')
        cols.insert(cols.index('OPERATOR_NAME'), 'FACILITY_NAME')
    dfSheet2_cleaned = dfSheet2_cleaned[cols]
    # cuối cùng phải còn lại là :
    # WDID	APP_ID	STATUS	FACILITY_NAME	OPERATOR_NAME	FACILITY_ADDRESS	FACILITY_CITY	FACILITY_STATE	FACILITY_ZIP	PRIMARY_SIC	SECONDARY_SIC	TERTIARY_SIC

    if st.checkbox("View Sheet2_2", key='BCB2'):
        st.write('(rows, cols) = ', len(dfSheet2_cleaned), len(dfSheet2_cleaned.columns))
        st.write(dfSheet2_cleaned)

    # Ghi cap nhat Sheet2
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet2_cleaned.to_excel(writer, sheet_name="Sheet2", index=False)
    F_excel_data_ao.seek(0)
    return F_excel_data_ao

# ham nay de In Sheet1, delete all rows duplicated and rows showing '4 56' in WDID
def Xli_P2_3(F_excel_data_ao):
    dfSheet1 = pd.read_excel(F_excel_data_ao, sheet_name='Sheet1')
    # drop all duplicated row
    dfSheet1_cleaned = dfSheet1.drop_duplicates()
    # xóa các dòng mà cột 'WDID' chứa chính xác chuỗi "4 56", còn lại các dòng ['WDID'] != '4 56'
    dfSheet1_cleaned = dfSheet1_cleaned[dfSheet1_cleaned['WDID'] != '4 56']
    # Delete all columns not in your tracker (columns A, J, K, U, X, and Y)
    dfSheet1_cleaned = dfSheet1_cleaned.drop(['PERMIT_TYPE', 'MONITORING_LATITUDE', 
        'MONITORING_LONGITUDE', 'ANALYTICAL_METHOD', 'DISCHARGE_END_DATE',	
        'DISCHARGE_END_TIME'],axis=1)
    #After deleting these columns, Sheet1 columns should look like the following: 
    # A – WDID; B – App ID; C – Status; D – Facility Name; E – Operator Name; F –Address; G – City; H – State; I – Zip; J – Primary SIC; K – Secondary SIC; L – Tertiary SIC
    # Sheet1 khong co STATUS, vay phai nhu ben phai day:  WDID	APP_ID	REPORTING_YEAR	REPORT_ID	EVENT_TYPE	MONITORING_LOCATION_NAME	MONITORING_LOCATION_TYPE	MONITOR_LOCATION_DESCRIPTION	SAMPLE_ID	SAMPLE_DATE	SAMPLE_TIME	DISCHARGE_START_DATE	DISCHARGE_START_TIME	PARAMETER	RESULT_QUALIFIER	RESULT	UNITS	MDL	RL	CERTIFIER_NAME	CERTIFIED_DATE
    if st.checkbox("View Sheet1_3", key='BCB3'):
        st.write('(rows, cols) = ', len(dfSheet1_cleaned), len(dfSheet1_cleaned.columns))
        st.write(dfSheet1_cleaned)

    # Ghi cap nhat Sheet2
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet1_cleaned.to_excel(writer, sheet_name="Sheet1", index=False)
    F_excel_data_ao.seek(0)
    return F_excel_data_ao

# Filter Sheet1 for only new sample data
def Xli_P2_4(F_excel_data_ao):
    dfSheet1 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet1")  # chứa cột APP_ID
    dfData = pd.read_excel(F_excel_data_ao, sheet_name="Data")  # chứa cột O và P

    # Giống =VLOOKUP(J2, Data!O:P, 2, FALSE)
    lookup_dict = pd.Series(dfData['PP'].values, index=dfData['OO']).to_dict()
    # Tìm vị trí (chỉ số) của cột 'APP_ID' trong Sheet1
    idx = dfSheet1.columns.get_loc('APP_ID')
    # Chèn cột mới 'VLOOKUP' vào trước 'APP_ID'
    dfSheet1.insert(loc=idx, column='VLOOKUP', value=dfSheet1['SAMPLE_DATE'].map(lookup_dict))
    # cột 'SAMPLE_DATE' là cột 'J' trong công thức =VLOOKUP(J2, Data!O:P, 2, FALSE

    # Xóa (lọc bỏ) tất cả các hàng có số trong cột 'VLOOKUP' 
    dfSheet1 = dfSheet1[~dfSheet1['VLOOKUP'].apply(lambda x: isinstance(x, (int, float)))]
    #dfSheet1 = dfSheet1[~dfSheet1['VLOOKUP'].apply(is_number_or_date)]
    # Giữ lại các dòng mà cột 'VLOOKUP' không chứa số trong Sêt1
    #dfSheet1 = dfSheet1[~dfSheet1['VLOOKUP'].astype(str).str.contains(r'\d', na=False)]
    #dfSheet1 = dfSheet1[~dfSheet1['VLOOKUP'].astype(str).str.contains(r'\d', na=False)]
    # sap xep theo 'VLOOKUP' tăng dần
    #dfSheet1 = dfSheet1.sort_values(by='VLOOKUP', ascending=False)
    dfSheet1 = dfSheet1.sort_values(by='VLOOKUP', ascending=True, na_position='first')

    if st.checkbox("View Sheet1_4", key='BCB4'):
        st.write('(rows, cols) = ', len(dfSheet1), len(dfSheet1.columns))
        st.write(dfSheet1)

    # Ghi cap nhat Sheet1
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet1.to_excel(writer, sheet_name="Sheet1", index=False)
    F_excel_data_ao.seek(0)
    return F_excel_data_ao

# DN Ham Check if facilities in Sheet1 are active
def Xli_P2_5(F_excel_data_ao):
    dfSheet1 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet1")  # chứa cột APP_ID
    dfSheet2 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet2")  # chứa cột O và P
    # Giống =VLOOKUP(C2,Sheet2!B:D,2,FALSE),
    lookup_dict = pd.Series(dfSheet2['STATUS'].values, index=dfSheet2['APP_ID']).to_dict()
    dfSheet1['VLOOKUP'] = dfSheet1['APP_ID'].map(lookup_dict)
    dfSheet1 = dfSheet1.sort_values(by='VLOOKUP', ascending=False)
    dfSheet1 = dfSheet1[dfSheet1['VLOOKUP'] == 'Active']

    if st.checkbox("View Sheet1_5", key='BCB5'):
        st.write('(rows, cols) = ', len(dfSheet1), len(dfSheet1.columns))
        st.write(dfSheet1)

    # Xóa (lọc bỏ) tất cả các hàng có số trong cột 'VLOOKUP' 
    #dfSheet1 = dfSheet1[~dfSheet1['VLOOKUP'].apply(lambda x: isinstance(x, (int, float)))]

    # Ghi cap nhat Sheet1
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet1.to_excel(writer, sheet_name="Sheet1", index=False)
    F_excel_data_ao.seek(0)
    return F_excel_data_ao

# DN Ham Choose the parameters to track in Sheet1
def Xli_P2_6(F_excel_data_ao):
    dfSheet1 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet1")  # chứa cột APP_ID
    #dfSheet2 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet2")  # chứa cột O và P

    # NEW tao menu de chon cac gia tri cua cot PARAMETER de lam viec 
    lpara = dfSheet1["PARAMETER"].dropna().unique().tolist()
    lgiulai = []
    popover = st.popover("Select parameter:")
    for pt in lpara:
        checked = popover.checkbox(pt, True)
        if checked:
            lgiulai.append(pt)
    # tai day list lgiulai da luu cac para chon

    # xem thu cac para da chon
    for pt in lgiulai:
        st.write(":red["+pt+"]") 

    # Xóa cột B
    dfSheet1 = dfSheet1.drop(dfSheet1.columns[1], axis=1)

    # Lọc giữ các dong co parameter mong muốn
    dfSheet1 = dfSheet1[dfSheet1["PARAMETER"].isin(lgiulai)]

    # Tách PARAMETER thành 2 cột PARAMETER và QUALIFIER
    split_cols = dfSheet1["PARAMETER"].str.split(",", n=1, expand=True)
    dfSheet1["PARAMETER"] = split_cols[0]
    param_index = dfSheet1.columns.get_loc("PARAMETER")
    dfSheet1.insert(param_index + 1, "QUALIFIER", split_cols[1])

    # xem Sheet1 da cap nhat
    if st.checkbox("View Sheet1_6", key='BCB6'):
        st.write('(rows, cols) = ', len(dfSheet1), len(dfSheet1.columns))
        st.write(dfSheet1)

    # Lưu lại Ghi cap nhat Sheet1
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet1.to_excel(writer, sheet_name="Sheet1", index=False)
    F_excel_data_ao.seek(0)
    return F_excel_data_ao

# DN Ham Make sure all the samples in Sheet1 are in 'mg/L' and not 'ug/L'
def Xli_P2_7(F_excel_data_ao):
    dfSheet1 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet1")  
    # Chuyển đổi từ ug/L sang mg/L
    mask_ug = dfSheet1["UNITS"] == "ug/L"
    dfSheet1.loc[mask_ug, "RESULT"] = dfSheet1.loc[mask_ug, "RESULT"] / 1000
    dfSheet1.loc[mask_ug, "UNITS"] = "mg/L"

    if st.checkbox("View Sheet1_7", key='BCB7'):
        st.write('(rows, cols) = ', len(dfSheet1), len(dfSheet1.columns))
        st.write(dfSheet1)

    # Lưu lại Ghi cap nhat Sheet1
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet1.to_excel(writer, sheet_name="Sheet1", index=False)
    F_excel_data_ao.seek(0)
    return F_excel_data_ao

def Xli_P2_8(F_excel_data_ao):
    dfSheet1 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet1")  
    dfSheet2 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet2")
    
    # Xoa cột C trong Sheet2
    dfSheet2 = dfSheet2.drop(dfSheet2.columns[2], axis=1)

    # Tạo bảng tra cứu giống vùng Sheet2 từ APP_ID
    lookup_cols = ["FACILITY_NAME", "OPERATOR_NAME", "FACILITY_ADDRESS", 
                "FACILITY_CITY", "FACILITY_STATE", "FACILITY_ZIP"]

    lookup_df = (
        dfSheet2
        .drop_duplicates(subset=["APP_ID"], keep="first")  # Giống VLOOKUP lấy bản ghi đầu tiên
        .set_index("APP_ID")[lookup_cols]
    )

    # Thêm 6 cột vào trước cột "Reporting Year" và map dữ liệu
        # Xác định vị trí cột Reporting Year
    pos = dfSheet1.columns.get_loc("REPORTING_YEAR")
        # Map từng cột lookup vào dfSheet1
    for i, col in enumerate(lookup_cols, start=1):
        dfSheet1.insert(pos + i - 1, col, dfSheet1["APP_ID"].map(lookup_df[col]))
        # Dòng tren này tương đương việc viết công thức VLOOKUP() và kéo sang 6 cột trong Excel.

    # Lưu lại Ghi cap nhat Sheet1
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet1.to_excel(writer, sheet_name="Sheet1", index=False)
        dfSheet2.to_excel(writer, sheet_name="Sheet2", index=False)
    F_excel_data_ao.seek(0)
    return F_excel_data_ao  
    
def Xli_P2_9(F_excel_data_ao):
    dfSheet1 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet1")  
    dfSheet2 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet2")
    #trong Sheet2 Xóa cột C → H, gexcePP_ID ngay ben trai PRIMARY_SIC
    # Xóa theo vị trí (C-H = cột 2 → 7 vì Python đếm từ 0)
    dfSheet2 = dfSheet2.drop(dfSheet2.columns[2:8], axis=1)
    # Đảm bảo cột sau khi xóa: WDID, APP_ID, PRIMARY_SIC, SECONDARY_SIC, TERTIARY_SIC

    # Trong Sheet1 – Thêm 3 cột mới ngay sau cột cuối cùng
        # Tạo bảng tra cứu từ Sheet2
    lookup_cols = ["PRIMARY_SIC", "SECONDARY_SIC", "TERTIARY_SIC"]

    lookup_df = (
        dfSheet2
        .drop_duplicates(subset=["APP_ID"], keep="first")  # giống VLOOKUP lấy bản ghi đầu tiên
        .set_index("APP_ID")[lookup_cols]
    )
        # Thêm 3 cột vào Sheet1 bằng map (tương đương viết công thức & kéo sang)
    for col in lookup_cols:
        dfSheet1[col] = dfSheet1["APP_ID"].map(lookup_df[col])

    # Xóa giá trị 0 trong TERTIARY_SIC và SECONDARY_SIC
    #Trong Excel, bước lọc “0” rồi Clear Contents thực chất là xóa tất cả giá trị bằng 0 trong cột.
        # Xóa giá trị 0 ở TERTIARY_SIC
    dfSheet1.loc[dfSheet1["TERTIARY_SIC"] == 0, "TERTIARY_SIC"] = None
        # Xóa giá trị 0 ở SECONDARY_SIC
    dfSheet1.loc[dfSheet1["SECONDARY_SIC"] == 0, "SECONDARY_SIC"] = None

    if st.checkbox("View Sheet1_9", key='BCB9_S1'):
        st.write('(rows, cols) = ', len(dfSheet1), len(dfSheet1.columns))
        st.write(dfSheet1)
    if st.checkbox("View Sheet2_9", key='BCB9_S2'):
        st.write('(rows, cols) = ', len(dfSheet2), len(dfSheet2.columns))
        st.write(dfSheet2)

    # Lưu lại Ghi cap nhat vao excel
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        dfSheet1.to_excel(writer, sheet_name="Sheet1", index=False)
        dfSheet2.to_excel(writer, sheet_name="Sheet2", index=False)
    F_excel_data_ao.seek(0)

    return F_excel_data_ao  

def Xli_P2_10(F_excel_data_ao):
    dfData = pd.read_excel(F_excel_data_ao, sheet_name="Data")  
    dfNew = pd.read_excel(F_excel_data_ao, sheet_name="Sheet1")
    # Điền "Old" vào cột OLD/NEW cho dữ liệu cũ trong sheet Data
        # Giả sử cột này tên là "OLD/NEW"
    dfData["OLD/NEW"] = "Old"

    # Gắn thêm dữ liệu mới từ Sheet1 vào Data
        # Bỏ hàng header trong Sheet1 (đã loại khi đọc file)
    df_combined = pd.concat([dfData, dfNew], ignore_index=True)

    # Thêm cột OLD/NEW = "New" cho dữ liệu mới
    # Xác định số hàng mới vừa thêm
    num_old = len(dfData)
    df_combined.loc[num_old:, "OLD/NEW"] = "New"

    # Lưu lại Ghi cap nhat vao excel
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_combined.to_excel(writer, sheet_name="Data", index=False)
    F_excel_data_ao.seek(0)

    # Xóa Sheet1 và Sheet2 khi xuất lại (chỉ giữ sheet "Data")
    #F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    #with pd.ExcelWriter(F_excel_data_ao) as writer:
    #    df_combined.to_excel(writer, sheet_name="Data", index=False)
    #F_excel_data_ao.seek(0)

    return F_excel_data_ao  

#---CAC HAM XLI PHAN III --------------------------------------
def Xli_P3_1234567(F_excel_data_ao):
    # Đọc dữ liệu từ file BytesIO() tên là F_excel_data_ao đã cập nhật ở bước cuối bên trên
    dfData = pd.read_excel(F_excel_data_ao, sheet_name="Data")  

    # Sort multi-level 
    dfData = dfData.sort_values(
        by=["OLD/NEW", "PARAMETER", "RESULT"],
        ascending=[True, True, False]
    )

    # Filter chỉ "New" 
    df_new = dfData[dfData["OLD/NEW"] == "New"].copy()
    # Danh sách facility quan tâm đặc biệt 
        # tao menu de chon cac gia tri cua cot FACILITY_NAME de lam viec 
    lpara = dfData["FACILITY_NAME"].dropna().unique().tolist()
    lgiulai2 = []
    popover = st.popover("Select FACILITY_NAME:")
    for pt in lpara:
        checked = popover.checkbox(pt, True)
        if checked:
            lgiulai2.append(pt)
    # tai day list lgiulai da luu cac para chon

    # xem thu cac para da chon
    for pt in lgiulai2: 
        st.write(':red['+pt+']') 
    # Giả sử bạn nhập tay hoặc lấy từ file khác
    special_facilities = lgiulai2 

    # ====== NAL, TNAL, NEL threshold ======
    NAL_annual = {
        "Aluminum": 0.75,
        "Ammonia": 2.14,
        "Arsenic": 0.15,
        "BOD": 30,
        "Cadmium": 0.0053,
        "COD": 120,
        "Copper": 0.0332,
        "Cyanide": 0.022,
        "Iron": 1.0,
        "Lead": 0.262,
        "Magnesium": 0.064,
        "Mercury": 0.0014,
        "Nickel": 1.02,
        "N+N": 0.68,
        "O&G": 15,
        "Phosphorus": 2.0,
        "Selenium": 0.005,
        "Silver": 0.0183,
        "TSS": 100,
        "Zinc": 0.26
    }

    NAL_instant = {
        "O&G": 25,
        "TSS": 400,
        # pH là trường hợp đặc biệt, sẽ xử lý riêng
    }

    TNAL_NEL_threshold = {
        "Ammonia": 4.7,
        "Cadmium": 0.0031,
        "Copper": 0.06749,
        "E. coli": 400,
        "Enterococci MPN": 104,
        "Fecal Coliform": 400,
        "Lead": 0.094,
        "Nitrate": 1.0,
        "Nitrite": 1.0,
        "N+N": 1.0,
        "Total Coliform": 10000,
        "Zinc": 0.159
    }

    # ====== Xác định exceedance ======
    def is_exceed(row):
        param = row["PARAMETER"]
        result = row["RESULT"]

        # pH special case
        if param == "pH":
            return result < 6.0 or result > 9.0

        # Check NAL annual
        if param in NAL_annual and result > NAL_annual[param]:
            return True

        # Check NAL instant
        if param in NAL_instant and result > NAL_instant[param]:
            return True

        # Check TNAL/NEL
        if param in TNAL_NEL_threshold and result > TNAL_NEL_threshold[param]:
            return True

        return False

    df_new["Exceedance"] = df_new.apply(is_exceed, axis=1)

    # Gắn cờ cho facilities cần xem xét 
    df_new["Special_Facility"] = df_new["FACILITY_NAME"].isin(special_facilities)
    #st.write(df_new)

    # Sort lại để xem 1 facility 
    df_facility_sorted = df_new.sort_values(
        by=["WDID", "REPORTING_YEAR", "PARAMETER", "RESULT"],
        ascending=[True, True, True, False]
    )

    styled_df = df_facility_sorted.style.apply(
        lambda row: ['background-color: yellow' if row['Exceedance'] else '' for _ in row],axis=1
        )

    if st.checkbox("View Sheet3_1", key='BCC1'):
        st.write('(rows, cols) = ', len(df_facility_sorted), len(df_facility_sorted.columns))
        st.write(df_facility_sorted)
    

    # Lưu lại Ghi cap nhat vao excel
    F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        styled_df.to_excel(writer, sheet_name="Data_new", index=False)
    F_excel_data_ao.seek(0)

    # Tạo nút tải xuống
    st.download_button(
        label="📥 Tải file Excel (Data_tracker_phan_III.xlsx)",
        data=F_excel_data_ao.getvalue(),
        file_name="Data_tracker_phan_III.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # chu y rang bay gio F_excel_data_ao co them sheet Data_new
    return F_excel_data_ao

def Xli_P3_8 (F_excel_data_ao):
    # Đọc Sheet3
    dfSheet3 = pd.read_excel(F_excel_data_ao, sheet_name="Sheet3")

    # 1️⃣ Đổi vị trí cột: đưa 'WDID' (cột B) lên trước 'App ID' (cột A)
    cols = dfSheet3.columns.tolist()
    # Giả sử 'App ID' là cols[0] và 'WDID' là cols[1]
    cols_reordered = [cols[1], cols[0]] + cols[2:]
    dfSheet3 = dfSheet3[cols_reordered]

    # 2️⃣ Xóa các cột L → AF (Question 4 Answer → Question TMDL Answer)
    dfSheet3 = dfSheet3.drop(columns=dfSheet3.loc[:, "QUESTION_4_ANSWER":"QUESTION_TMDL_ANSWER"].columns)

    # 3️⃣ Xóa các cột E → I (Region → Question 2 Explanation)
    dfSheet3 = dfSheet3.drop(columns=dfSheet3.loc[:, "REGION":"QUESTION_2_EXPLANATION"].columns)

    # 4️⃣ Đảm bảo thứ tự cột đúng như mong muốn
    desired_order = [
        "WDID",
        "APP_ID",
        "REPORT_ID",
        "REPORT_YEAR",
        "QUESTION_3_ANSWER",
        "QUESTION_3_EXPLANATION"
    ]
    dfSheet3 = dfSheet3[desired_order]

    # ==== 6. Xóa các hàng có Question 3 Answer == "N" hoặc rỗng ====
    dfSheet3 = dfSheet3[~dfSheet3["QUESTION_3_ANSWER"].isin(["N", None, ""])]

    # ==== 7. Chèn cột trống trước cột "App ID" ====
    dfSheet3.insert(1, "App_ID_from_Data", None)  # cột mới B

    # ==== 8. Thực hiện tra cứu (VLOOKUP) từ sheet "Data" ====
    # Giả sử sheet "Data" có cột O và P
    #data_df = pd.read_excel("input.xlsx", sheet_name="Data", usecols="O:P")
    dfData = pd.read_excel(F_excel_data_ao, sheet_name="Data", usecols="O:P")

    # đặt lại tên cột của dfData thành đúng 2 tên mới: 
    dfData.columns = ["lookup_key", "lookup_value"]

    # Tạo cột mới dựa vào J2 = 'REPORT_YEAR' (hoặc thay 'J' bằng cột thực tế)
    dfSheet3["App_ID_from_Data"] = dfSheet3["REPORT_YEAR"].map(
        dict(zip(dfData["lookup_key"], dfData["lookup_value"]))
    )

    if st.checkbox("Edit Sheet3_2", key='BCC2'):
        st.write('(rows, cols) = ', len(dfSheet3), len(dfSheet3.columns))
        edited_dfSheet3 = st.data_editor(dfSheet3)

        # Lưu lại Ghi cap nhat vao excel
        F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
        with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            edited_dfSheet3.to_excel(writer, sheet_name="Sheet3_new", index=False)
        F_excel_data_ao.seek(0)
    else:
        F_excel_data_ao.seek(0)  # quan trọng: để writer đọc được file hiện tại
        with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            dfSheet3.to_excel(writer, sheet_name="Sheet3_new", index=False)
        F_excel_data_ao.seek(0)


    return F_excel_data_ao


#-------------------
def ThucThiPhan_1():
    # cac vung de chon
    regions = st.selectbox("Select a Region:", 
                ("Region 1 - North Coast",
                "Region 2 - San Francisco Bay",
                "Region 3 - Central Coast",
                "Region 4 - Los Angeles",
                "Region 5F - Fresno",
                "Region 5R - Redding",
                "Region 5S - Sacramento",
                "Region 6A - South Lake Tahoe",
                "Region 6B - Victorville",
                "Region 7 - Colorado River Basin",
                "Region 8 - Santa Ana",
                "Region 9 - San Diego"),
                index=None,
                placeholder="No selected Region",
                )
    #neu mot vung duoc chon thi lam
    LOI='OK'
    if regions:
        placeholder_1 = st.empty()
        placeholder_1.write('Wait for downloading 2 files of ' + regions)
        #thuc thi ham download_data_smarts(regions) va tra ve list cac file da tai 
        try :
            lfile_datai = download_data_smarts(regions)
            placeholder_1.write('Downloaded files:')
            st.write(lfile_datai)
        except:
            LOI='LOI'
            placeholder_1.write('Tai file không đạt!')
    if LOI == 'LOI':
        st.write('Nếu không đạt, '+ ':red[ mở trực tiếp trang sau làm theo các bước để tải:]')
        st.markdown("1-[Open Page SMARTS](https://smarts.waterboards.ca.gov/smarts/SwPublicUserMenu.xhtml)", unsafe_allow_html=True)
        st.write('2-Click on “Download NOI Data By Regional Board”')
        st.write('3-Select your region from the dropdown menu')
        st.write('4-Click on both “Industrial Application Specific Data” and “Industrial Ad Hoc Reports - Parameter Data”')
        st.write('5-Data will be downloaded to two separate .txt files, each titled “file”')
        st.write('6-Nên đổi tên 2 file thành Industrial_Application_Specific_Data và Industrial_Ad_Hoc_Reports_-_Parameter_Data rồi chép vào thư mục riêng của bạn để dễ làm việc ở các bước sau.')

def update_checkbox_sidebar(tep_mo):
    with open(tep_mo, "r", encoding="utf-8") as f:
        data = json.load(f)   
    # Ghép key và value thành chuỗi "key - value"
    if '1' in tep_mo:
        options1 = [f"{k} - {v}" for k, v in data.items()]
        for op in options1:
            if '#' not in op: 
                st.write(op)
    elif '2' in tep_mo:
        options2 = [f"{k} - {v}" for k, v in data.items()]
        for op in options2:
            if '#' not in op: 
                st.write(op)
    elif '3' in tep_mo:
        options3 = [f"{k} - {v}" for k, v in data.items()]
        for op in options3:
            if '#' not in op: 
                st.write(op)
    else:
        options4 = [f"{k} - {v}" for k, v in data.items()]
        for op in options4:
            if '#' not in op: 
                st.write(op)

def Xem_do_thi_1():
    # Example data
    data = {
        "Month": ["Jan", "Feb", "Mar", "Apr", "May", "Jun"],
        "Sales": [150, 200, 250, 220, 300, 400],
        "Expenses": [100, 120, 180, 160, 210, 280]
    }
    # Create DataFrame
    df = pd.DataFrame(data)
    # Set index for better x-axis labels
    df.set_index("Month", inplace=True)
    st.title("📈 Company Sales and Expenses Over Months")
    # Line chart
    st.line_chart(df)


def Xem_do_thi_2():
    # Example data
    data = {
        "Product": ["A", "B", "C", "D"],
        "Sales": [300, 450, 150, 500],
        "Profit": [80, 120, 300, 20]
    }
    # Create DataFrame
    df = pd.DataFrame(data)
    # Set 'Product' as index (so it appears on x-axis)
    df.set_index("Product", inplace=True)
    st.title("📊 Sales and Profit by Product")
    # Streamlit bar chart
    st.bar_chart(df)

def Xem_do_thi_3():
    data = {
        "Product": ["A", "B", "C", "D"],
        "Sales": [300, 450, 150, 500],
        "Profit": [80, 120, 50, 200]
    }

    # Create DataFrame
    df = pd.DataFrame(data)
    df.set_index("Product", inplace=True)

    st.title("📊 Horizontal Bar Plot Example")

    # Create matplotlib horizontal bar plot
    fig, ax = plt.subplots()
    df.plot(kind="barh", ax=ax)

    # Customize
    ax.set_title("Sales and Profit by Product")
    ax.set_xlabel("Amount ($)")
    ax.set_ylabel("Product")

    # Show in Streamlit
    st.pyplot(fig)


def Xem_do_thi_4():
    # Giả lập dữ liệu: thu nhập của 1000 nhân viên (triệu VND)
    np.random.seed(42)
    incomes = np.random.normal(loc=15, scale=5, size=1000)  # trung bình 15, lệch chuẩn 5
    incomes = np.clip(incomes, 5, 50)  # Giới hạn từ 5 đến 50 triệu

    df = pd.DataFrame({"Income": incomes})

    # Vẽ histogram
    fig, ax = plt.subplots()
    df["Income"].hist(bins=20, ax=ax, edgecolor="black")
    ax.set_title("Biểu đồ Histogram về Phân bố thu nhập nhân viên")
    ax.set_xlabel("Thu nhập (triệu VND)")
    ax.set_ylabel("Số lượng nhân viên")
    st.pyplot(fig)

def Xem_do_thi_5():
    st.title("📦 Phân bố điểm thi của học sinh 3 lóp A,B,C bằng Box Plot")
    # Tạo dữ liệu giả lập
    np.random.seed(42)
    data = {
        "Class": (["A"] * 30) + (["B"] * 30) + (["C"] * 30),
        "Score": list(np.random.normal(75, 10, 30)) +   # Lớp A: trung bình 75, lệch chuẩn 10
                list(np.random.normal(65, 15, 30)) +   # Lớp B: trung bình 65, lệch chuẩn 15
                list(np.random.normal(80, 50, 30))      # Lớp C: trung bình 80, lệch chuẩn 5
    }

    df = pd.DataFrame(data)
    # Vẽ box plot
    fig, ax = plt.subplots()
    df.boxplot(column="Score", by="Class", ax=ax)
    # Tùy chỉnh
    ax.set_title("So sánh phân bố điểm thi giữa các lớp")
    ax.set_xlabel("Lớp học")
    ax.set_ylabel("Điểm số")
    plt.suptitle("")  # Xóa tiêu đề mặc định của pandas

    # Hiển thị trong Streamlit
    st.pyplot(fig)
    
    st.markdown("""
    ✅ Ở ví dụ này:
    Lớp A có điểm khá ổn định quanh 75.
    Lớp B phân bố rộng, nhiều học sinh chênh lệch.
    Lớp C tập trung quanh 80, ít biến động.
    👉 Đây chính là tình huống điển hình mà chỉ box plot mới diễn tả được, 
    còn bar chart chỉ cho bạn thấy trung bình, mất hết thông tin về phân bố.
    """)
    st.markdown("""
    ### 📌 Vì sao dùng Box Plot?
    - Hiển thị **median (trung vị)**: mức điển hình của lớp.
    - Cho thấy **khoảng tứ phân vị (IQR)**: độ phân tán.
    - Thấy ngay **outliers (điểm bất thường)**, ví dụ học sinh điểm quá thấp hoặc quá cao.
    - Các biểu đồ khác như **bar chart, line chart** không thể hiện được những thông tin này. 
    [Read more here](http://sociologyhue.edu.vn/blog/post/22288)
    """)

def Xem_do_thi_6():
    st.title("🌈 Area Plot Example")
    # Example dataset
    data = {
        "Month": ["Jan", "Feb", "Mar", "Apr", "May", "Jun"],
        "Sales": [150, 200, 250, 220, 300, 400],
        "Expenses": [100, 120, 180, 160, 210, 280]
    }
    # Create DataFrame
    df = pd.DataFrame(data)
    df.set_index("Month", inplace=True)

    # Create matplotlib figure
    fig, ax = plt.subplots()
    df.plot(kind="area", alpha=0.5, ax=ax)  # alpha để trong suốt nhìn rõ chồng lên nhau

    # Customize
    ax.set_title("Sales vs Expenses (Area Plot)")
    ax.set_xlabel("Month")
    ax.set_ylabel("Amount ($)")
    ax.grid(True)

    # Show in Streamlit
    st.pyplot(fig)

def Xem_do_thi_7():
    st.title("🥧 Pie Chart Example")

    # Example dataset
    data = {
        "Product": ["A", "B", "C", "D"],
        "Sales": [300, 450, 150, 500]
    }

    # Create DataFrame
    df = pd.DataFrame(data)

    # Create matplotlib figure
    fig, ax = plt.subplots()
    ax.pie(df["Sales"], labels=df["Product"], autopct="%1.1f%%", startangle=90)
    ax.set_title("Sales Distribution by Product")

    # Show in Streamlit
    st.pyplot(fig)

def Xem_do_thi_8():
    st.title("🔹 Scatter Plot Example")
    # Example dataset
    np.random.seed(42)
    data = {
        "Advertising": np.random.randint(50, 200, 20),  # Chi phí quảng cáo
        "Sales": np.random.randint(100, 500, 20)        # Doanh số
    }

    # Create DataFrame
    df = pd.DataFrame(data)

    # Create matplotlib figure
    fig, ax = plt.subplots()
    ax.scatter(df["Advertising"], df["Sales"], color="blue", s=100, alpha=0.7, edgecolors="k")

    # Customize
    ax.set_title("Sales vs Advertising")
    ax.set_xlabel("Advertising ($)")
    ax.set_ylabel("Sales ($)")
    ax.grid(True)

    # Show in Streamlit
    st.pyplot(fig)

def Xem_do_thi_9():
    st.title("🔷 Hexbin Plot Example")
    # Tạo dữ liệu ví dụ
    np.random.seed(42)
    x = np.random.randn(1000) * 50 + 200   # Dữ liệu Advertising
    y = np.random.randn(1000) * 80 + 300   # Dữ liệu Sales

    df = pd.DataFrame({"Advertising": x, "Sales": y})

    # Tạo figure
    fig, ax = plt.subplots(figsize=(7,5))

    # Vẽ hexbin plot
    hb = ax.hexbin(df["Advertising"], df["Sales"], gridsize=30, cmap="Blues", mincnt=1)

    # Thêm colorbar
    cb = fig.colorbar(hb, ax=ax)
    cb.set_label("Number of points")

    # Tùy chỉnh
    ax.set_title("Sales vs Advertising (Hexbin Plot)")
    ax.set_xlabel("Advertising ($)")
    ax.set_ylabel("Sales ($)")
    ax.grid(True)

    # Hiển thị trong Streamlit
    st.pyplot(fig)
                

#========================= MAIN =====================================================================
# global
DACO_EXCEL_4SHEET=False
PHANIIFINISHED=False
PHANIIIFINISHED=False
tepexcel_ao=None
# TIEU DE APP
#st.header('🔎 Consider CA Rai')
st.title('🔎 Review Reported Wastewater Data')
#-----------------------------------------

# Phan sider ben trai ---------------------------------------------------------------------------
with st.sidebar:
    st.header('🏷️ :red[LOOK UP DOCUMENT]')
    # Xem tai lieu SMARTS
    st.write("---")
    checkbox_sidebar_0 = st.checkbox(":blue[📌 SMARTS documents used as a basis for writing this program]", key='PL0', value=False)
    if checkbox_sidebar_0:
        # Đọc nội dung file Markdown
        with open("hd-lam-app-cho-thong.md", "r", encoding="utf-8") as f:
            md_content = f.read()
        st.markdown(md_content, unsafe_allow_html=True)
    
    st.write("---")
    checkbox_sidebar_1 = st.checkbox(":green[📌 Headers in Sheet1 (Industrial_Ad_Hoc_Reports)]", key='PL1', value=False)
    if checkbox_sidebar_1:
        tep_1 = "Headers/dict_sheet1.json"
        update_checkbox_sidebar(tep_1)

    st.write("---")
    # Xem Header Sheet2 
    checkbox_sidebar_2 = st.checkbox(":green[📌 Headers in Sheet2 (Industrial_Application_Specific_Data)]", key='PL2', value=False)
    if checkbox_sidebar_2:
        tep_2 = "Headers/dict_sheet2.json"
        update_checkbox_sidebar(tep_2)


    st.write("---")
    # Xem Header Sheet3 
    checkbox_sidebar_3 = st.checkbox(":green[📌 Headers in Sheet3 (Industrial_Annual_Reports)]", key='PL3', value=False)
    if checkbox_sidebar_3:
        tep_3 = "Headers/dict_sheet3.json"
        update_checkbox_sidebar(tep_3)

    st.write("---")
    # Xem Header Data
    checkbox_sidebar_4 = st.checkbox(":green[📌 Headers in Data]", key='PL4', value=False)
    if checkbox_sidebar_4:
        tep_4 = "Headers/dict_data.json"
        update_checkbox_sidebar(tep_4)


    st.write("---")
    # Minh hoa vai loai do thi
    checkbox_sidebar_5 = st.checkbox(":red[📌Illustrate data examples]", key='PL5', value=False)
    if checkbox_sidebar_5:
        loai_do_thi = st.radio(
            "Chon loai do thi",
            ["Line plot",
            "Bar plot", 
            "Barh plot", 
            "Histogram", 
            "Box plot", 
            "Area plot", 
            "Pie chart",
            "Scatter plot",
            "Hexbin plot"
            ],
            index=None,
        )
        if loai_do_thi == "Line plot":
            Xem_do_thi_1()
        elif loai_do_thi == "Bar plot":
            Xem_do_thi_2()
        elif loai_do_thi == "Barh plot":
            Xem_do_thi_3()
        elif loai_do_thi == "Histogram":
            Xem_do_thi_4()
        elif loai_do_thi == "Box plot":
            Xem_do_thi_5()
        elif loai_do_thi == "Area plot":
            Xem_do_thi_6()
        elif loai_do_thi == "Pie chart":
            Xem_do_thi_7()
        elif loai_do_thi == "Scatter plot":
            Xem_do_thi_8()
        elif loai_do_thi == "Hexbin plot":
            Xem_do_thi_9()



# I. TAI FILES TXT DU LIEU VE TU -----------------------------------------------------------------
st.header('✅ I. Download the data', divider=True)
ThucThiPhan_1()

# II Them data moi vao trinh theo doi -------------------------------------------------------------
st.header('✅ II. Add the new data to your tracker', divider=True)

laydatafrom = st.radio(
    "WHERE GET DATA ", 
    [":blue[From Local]",":green[From Server]", ":red[Empty]"],
    index=2,horizontal=True , label_visibility="visible") 

if laydatafrom==":red[Empty]":
    DACO_EXCEL_4SHEET=False
    pass  

elif laydatafrom==":blue[From Local]":
    # Add the new data to your tracker 
    # - Upload 4 files
    uploaded_files = st.file_uploader(
        'Upload 1 lần 4 files: "...Industrial_Ad_Hoc...", "...Industrial_Application...", "...Industrial_Annual_Reports.txt...", "...Data_Tracker..." ' + ' :red[(nên đặt 4 files này liền nhau trong 1 thư mục)]',
        type=['txt', 'xlsx'],  
        accept_multiple_files=True
    )
    st.write(len(uploaded_files),'files')
    for file in uploaded_files:
        st.write(file.name)

    if uploaded_files and len(uploaded_files) == 4:
        # Phân loại file theo đuôi và tên
        uploaded_f1 = next((f for f in uploaded_files if "industrial_ad_hoc" in f.name.lower()), None)
        uploaded_f2 = next((f for f in uploaded_files if "industrial_application" in f.name.lower()), None)
        uploaded_f3 = next((f for f in uploaded_files if "industrial_annual_reports" in f.name.lower()), None)
        uploaded_f4 = next((f for f in uploaded_files if f.name.lower().endswith(".xlsx")), None)

        if uploaded_f1 and uploaded_f2 and uploaded_f3 and uploaded_f4:
            try:
                #df = pd.read_csv("file.csv", on_bad_lines="skip")
                df1 = pd.read_csv(uploaded_f1, sep='\t', encoding='cp1252', on_bad_lines="skip")
                df2 = pd.read_csv(uploaded_f2, sep='\t', encoding='cp1252', on_bad_lines="skip")
                df3 = pd.read_csv(uploaded_f3, sep='\t', encoding='cp1252', on_bad_lines="skip")
                #dfData = pd.read_excel(uploaded_f4, sheet_name="Data")  # Chỉ đọc sheet "Data"
            except Exception as e:
                st.error(f"⚠️ Lỗi khi đọc file: {e}")
                st.stop()
            #---
            # Đọc file Excel đã upload
            excel_data = uploaded_f4.read()

            # Ghi DataFrame TXT vào file Excel đã upload
            F_excel_data_ao = BytesIO()
            with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl") as writer:
                # Ghi lại các sheet cũ của file Excel gốc
                original_excel = pd.ExcelFile(BytesIO(excel_data))
                for sheet_name in original_excel.sheet_names:
                    df_old = pd.read_excel(original_excel, sheet_name=sheet_name)
                    df_old.to_excel(writer, sheet_name=sheet_name, index=False)
                # Thêm / Ghi đè sheet "Sheet1" bằng dữ liệu từ file TXT
                df1.to_excel(writer, sheet_name="Sheet1", index=False)
                df2.to_excel(writer, sheet_name="Sheet2", index=False)
                df3.to_excel(writer, sheet_name="Sheet3", index=False)

            # 3. Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            DACO_EXCEL_4SHEET=True
else:
    # Đường dẫn thư mục TAM (nằm ngang với streamlit_app.py)
    BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Datatest")

    # Tên file mong muốn
    file_names = ["Region_1_-_North_Coast_-_Industrial_Ad_Hoc_Reports_-_Parameter_Data.txt",
            "Region_1_-_North_Coast_-_Industrial_Application_Specific_Data.txt", 
            "Region_1_-_North_Coast_-_Industrial_Annual_Reports.txt",
            "Data_Tracker_X.xlsx"
            ]
    # Danh sách đường dẫn file trên server
    server_files = [os.path.join(BASE_DIR, name) for name in file_names]
    #server_files = file_names
    # Kiểm tra xem tất cả file có sẵn trên server không
    if all(os.path.exists(path) for path in server_files):
        st.info("📂 Đang dùng file trong thư mục Datatest.")
            #with open(server_files[0], "r", encoding="utf-8") as f1:
            #    txt_content_f1 = f1.read()
            #with open(server_files[1], "r", encoding="utf-8") as f2:
            #    txt_content_f2 = f2.read()
        df1 = pd.read_csv(server_files[0], sep='\t', encoding='cp1252')
        df2 = pd.read_csv(server_files[1], sep='\t', encoding='cp1252')
        df3 = pd.read_csv(server_files[2], sep='\t', encoding='cp1252')
        #dfData = pd.read_excel(server_files[3])
        excel_data = pd.read_excel(server_files[3])

    else:
        st.warning("📤 File chưa có trong Datatest, vui lòng upload.")
        exit()
    #excel_data = pd.read_excel(server_files[3])

    # Lưu lại Ghi cap nhat vao excel
    F_excel_data_ao = BytesIO()
    with pd.ExcelWriter(F_excel_data_ao, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
        df3.to_excel(writer, sheet_name="Sheet3", index=False)
        excel_data.to_excel(writer, sheet_name="Data", index=False)

    # Tạo nút tải xuống
    st.download_button(
        label="📥 Tải file Excel (Data_tracker_goc.xlsx)",
        data=F_excel_data_ao.getvalue(),
        file_name="Data_tracker_goc.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    DACO_EXCEL_4SHEET=True

if DACO_EXCEL_4SHEET==True :
    
    #--------------------------------------
    #st.subheader(":red[➡️ Add the new data to your tracker]")

    #checkbox1 = st.checkbox("📌:blue[1. Xóa các dòng mà STATUS ≠ 'Active' trong các dòng có APP_ID trùng lặp in Sheet2]", key='CB1')
    checkbox1 = st.checkbox("📌:blue[1. In your data_tracker.xlsx, create Sheet1, Sheet2, Sheet3 contain 3 file.txt]", key='CB1', value=True)
    if checkbox1:
        # tra ve kq la file ao da update cung ten F_excel_data_ao 
        F_excel_data_ao = Xli_P2_1(F_excel_data_ao) 
        if F_excel_data_ao:
            st.write(':green[Xli_P2_1 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_1.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_1.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    #checkbox2 = st.checkbox("📌:blue[2. Delete, move, re-order columns in Sheet2]", key='CB2')
    checkbox2 = st.checkbox("📌:blue[2. Get Sheet2 into the proper format for your tracker]", key='CB2', value=True)
    if checkbox1 and checkbox2:
        F_excel_data_ao = Xli_P2_2(F_excel_data_ao)
        if F_excel_data_ao:
            st.write(':green[Xli_P2_2 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_2.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_2.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    #checkbox3 = st.checkbox("📌:blue[3. Delete all rows duplicated and rows showing '4 56' in WDID in Sheet1]", key='CB3')
    checkbox3 = st.checkbox("📌:blue[3. Get Sheet1 into the proper format for your tracker]", key='CB3', value=True)
    if checkbox1 and checkbox2 and checkbox3:
        F_excel_data_ao = Xli_P2_3(F_excel_data_ao)
        if F_excel_data_ao:
            st.write(':green[Xli_P2_3 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_2.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_3.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    checkbox4 = st.checkbox("📌:blue[4. Filter Sheet1 for only new sample data]", key='CB4', value=True)
    if checkbox1 and checkbox2 and checkbox3 and checkbox4:
        F_excel_data_ao = Xli_P2_4(F_excel_data_ao)
        if F_excel_data_ao:
            st.write(':green[Xli_P2_4 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_4.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_4.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    checkbox5 = st.checkbox("📌:blue[5. Check if facilities in Sheet1 are active]", key='CB5', value=True)
    if checkbox1 and checkbox2 and checkbox3 and checkbox4 and checkbox5:
        F_excel_data_ao = Xli_P2_5(F_excel_data_ao)
        if F_excel_data_ao:
            st.write(':green[Xli_P2_5 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_5.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_5.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    checkbox6 = st.checkbox("📌:blue[6. Choose the parameters to track in Sheet1]", key='CB6', value=True)
    if checkbox1 and checkbox2 and checkbox3 and checkbox4  and checkbox5 and checkbox6:
        F_excel_data_ao = Xli_P2_6(F_excel_data_ao)

        if F_excel_data_ao:
            st.write(':green[Xli_P2_6 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_6.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_6.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


    checkbox7 = st.checkbox("📌:blue[7. Make sure all the samples in Sheet1 are in mg/L and not ug/L]", key='CB7', value=True)
    if checkbox1 and checkbox2 and checkbox3 and checkbox4  and checkbox5  and checkbox6  and checkbox7:
        F_excel_data_ao = Xli_P2_7(F_excel_data_ao)

        if F_excel_data_ao:
            st.write(':green[Xli_P2_7 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_7.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_7.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


    checkbox8 = st.checkbox("📌:blue[8. Add facility information from Sheet2 into Sheet1]", key='CB8', value=True)
    if checkbox1 and checkbox2 and checkbox3 and checkbox4  and checkbox5  and checkbox6  and checkbox7 and checkbox8:
        F_excel_data_ao = Xli_P2_8(F_excel_data_ao)

        if F_excel_data_ao:
            st.write(':green[Xli_P2_8 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_8.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_8.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    checkbox9 = st.checkbox("📌:blue[9. Add in SIC Codes from Sheet2 into Sheet1]", key='CB9', value=True)
    if checkbox1 and checkbox2 and checkbox3 and checkbox4  and checkbox5  and checkbox6  and checkbox7 and checkbox8  and checkbox9:
        F_excel_data_ao = Xli_P2_9(F_excel_data_ao)

        if F_excel_data_ao:
            st.write(':green[Xli_P2_9 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_9.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_9.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    checkbox10 = st.checkbox("📌:blue[10. Combine new data from Sheet1 into existing Data tracker]", key='CB10', value=True)
    if checkbox1 and checkbox2 and checkbox3 and checkbox4  and checkbox5  and checkbox6  and checkbox7 and checkbox8  and checkbox9:
        F_excel_data_ao = Xli_P2_10(F_excel_data_ao)

        if F_excel_data_ao:
            st.write(':green[Xli_P2_10 finished.]')
            # Tạo nút tải xuống
            st.download_button(
                label="📥 Tải file Excel (Data_tracker_add2sheet_10.xlsx)",
                data=F_excel_data_ao.getvalue(),
                file_name="Data_tracker_add2sheet_10.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        PHANIIFINISHED=True

# III Them data moi vao trinh theo doi -------------------------------------------------------------
st.header('✅ III. Analyze the new data', divider=True)
if PHANIIFINISHED:
    
    checkboxIII_1234567 = st.checkbox("📌:blue[1. Xem cơ sở nào có các chỉ số vượt ngưỡng ]", key='CB31', value=True)
    if checkboxIII_1234567:
        F_excel_data_ao = Xli_P3_1234567(F_excel_data_ao)
        
    if F_excel_data_ao:
        st.write(':green[Xli_P3_1234567 finished.]')
        # Tạo nút tải xuống
        st.download_button(
            label="📥 Tải file Excel (Data_tracker_add3sheet_1234567.xlsx)",
            data=F_excel_data_ao.getvalue(),
            file_name="Data_tracker_add3sheet_1234567.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    # Phan sau lay lai sheet Data de lam viec (KHONG lay Data_new)
    #-------------------------------
    # 8. Kiểm tra xem cơ sở nào đang nói dối trong báo cáo hàng năm 
    # về việc lấy mẫu tất cả các QSE

    checkboxIII_8 = st.checkbox("📌:blue[2. Check to see which facilities are lying in annual reports about sampling all QSEs]", key='CB32', value=True)
    if checkboxIII_8:
        F_excel_data_ao = Xli_P3_8(F_excel_data_ao)

    if F_excel_data_ao:
        st.write(':green[Xli_P3_8 finished.]')

        # Tạo nút tải xuống
        st.download_button(
            label="📥 Tải file Excel (Data_tracker_add3sheet_8.xlsx)",
            data=F_excel_data_ao.getvalue(),
            file_name="Data_tracker_add3sheet_8.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        PHANIIIFINISHED=True







# IV Them data moi vao trinh theo doi -------------------------------------------------------------
st.header('✅ IV. Visualize the data', divider=True)
if PHANIIIFINISHED:
    def pivot_summary(df, index_cols, value_col="Result", agg="mean"):
        """Tạo pivot table từ dataframe"""
        pivot = pd.pivot_table(
            df,
            index=index_cols,
            columns="Parameter",
            values=value_col,
            aggfunc=agg,
            fill_value=0
        ).reset_index()
        return pivot

    def visualize_data():
        st.subheader("📊 Water Data Summary Tool")

        # Upload Excel
        uploaded = st.file_uploader("Upload Excel file", type=["xlsx"])
        if uploaded is not None:
            df = pd.read_excel(uploaded, sheet_name="Data")
            st.success("✅ File loaded")

            # Làm sạch dữ liệu
            df["Result"] = df["Result"].fillna(0)

            # Chọn index cho pivot
            index_options = ["Facility", "MonitoringLocation", "ReportingYear"]
            index_cols = st.multiselect("Chọn các cột làm Rows:", index_options, default=["ReportingYear"])

            # Chọn phép tính
            agg_func = st.selectbox("Chọn phép tính:", ["mean", "count", "sum"])

            if st.button("Tạo Pivot Table"):
                pivot = pivot_summary(df, index_cols, "Result", agg=agg_func)
                st.dataframe(pivot)

                # Xuất Excel
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                    pivot.to_excel(writer, sheet_name="Summary", index=False)
                st.download_button("📥 Download Excel", data=output.getvalue(),
                                file_name="summary.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # Vẽ biểu đồ
                if "ReportingYear" in pivot.columns:
                    for param in df["Parameter"].unique():
                        if param in pivot.columns:
                            fig, ax = plt.subplots()
                            pivot.groupby("ReportingYear")[param].mean().plot(kind="bar", ax=ax)
                            ax.set_title(f"{agg_func.capitalize()} of {param} by Year")
                            st.pyplot(fig)

    visualize_data()

end_time = time.time()
elapsed_time = (end_time - start_time)/60

st.write(f"⏳ It took : {elapsed_time:.2f} minutes")